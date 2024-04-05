import re
import random
import time
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment

MAX_PAGES = 10 
PAGE_LOAD_TIMEOUT = 20000  
SCROLL_PAUSE_MIN = 2  
SCROLL_PAUSE_MAX = 4  
DATE_PATTERN = r'\d{1,2}\s(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s\d{4}'

# def get_detail_product(target_url):

#     url = target_url
#     detail_page = requests.get(url)
#     detail_soup = BeautifulSoup(detail_page.content, 'html.parser')

#     specification_title = detail_soup.find_all("dt", class_="ux-labels-values__labels")
#     specification_value = detail_soup.find_all("dd", class_="ux-labels-values__values")
    
#     product_details = []
#     for index, value in enumerate(specification_title):
#         product_details.append((value.text, specification_value[index].text))

#     return product_details

def scrape_and_save(keyword):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headings = ["Product Name", "Price", "Sold Date", "Product URL"]
    for idx, heading in enumerate(headings, start=1):
        cell = ws.cell(row=1, column=idx, value=heading)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

        row = 2
        for pgn_no in range(1, MAX_PAGES + 1):
            try:
                page = context.new_page()

                for attempt in range(3): 
                    try:
                        page.goto(f'https://www.ebay.co.uk/sch/i.html?_stpos=rm26lu&_nkw={keyword}&LH_PrefLoc=1&LH_Complete=1&_fcid=3&LH_Sold=1&_ipg=240&_pgn={pgn_no}', timeout=PAGE_LOAD_TIMEOUT)
                        break 
                    except Exception as e:
                        print(f"Attempt {attempt+1}: Error loading page {pgn_no}: {e}")
                        if attempt < 2:
                            print("Retrying...")
                        else:
                            print("Max attempts reached. Skipping page.")
                            continue

                page.wait_for_selector('.s-item')

                last_height = page.evaluate("() => document.body.scrollHeight")

                while True:
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(random.uniform(SCROLL_PAUSE_MIN, SCROLL_PAUSE_MAX))

                    new_height = page.evaluate("() => document.body.scrollHeight")
                    if new_height == last_height:
                        break
                    last_height = new_height

                results = page.query_selector_all('.s-item')

                for result in results:
                    try:
                        title_elem = result.query_selector('.s-item__title')
                        price_elem = result.query_selector('.s-item__price')
                        product_url = result.query_selector('.s-item__link').get_attribute('href')

                        try:
                            sold_date_elem = result.query_selector('.s-item__title--tag')
                            sold_date = sold_date_elem.inner_text()
                            matches = re.findall(DATE_PATTERN, sold_date)
                            if matches:
                                sold_date = matches[0]
                            else:
                                sold_date = ''
                        except:
                            sold_date = "N/A"

                        title = title_elem.inner_text()
                        price = price_elem.inner_text()

                        ws.cell(row=row, column=1, value=title)
                        ws.cell(row=row, column=2, value=price)
                        ws.cell(row=row, column=3, value=sold_date)
                        ws.cell(row=row, column=4, value=product_url)
                        row += 1
                    except Exception as e:
                        print(f"Error processing result: {e}")

                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

                print(f'Page {pgn_no}/{MAX_PAGES} scraped successfully')

            except Exception as err:
                print(f"Error scraping page {pgn_no}: {err}")
                break

        wb.save(f"{keyword}_results.xlsx")
        print("Scraping and saving complete.")

if __name__ == "__main__":
    keyword = input("Enter a keyword to search on eBay: ")
    scrape_and_save(keyword)
