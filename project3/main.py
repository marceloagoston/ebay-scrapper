from bs4 import BeautifulSoup
import requests
import openpyxl

from datetime import datetime

def read_urls_from_excel(file_path):
    urls = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.startswith('http'):
                    urls.append(cell)
                    
        workbook.close()
    except Exception as e:
        print(f"Error reading Excel file: {e}")

    return urls

def get_detail_product(target_url):

    print('** Request to: ', target_url)
    detail_page = requests.get(target_url)
    detail_soup = BeautifulSoup(detail_page.content, 'html.parser')

    title = detail_soup.find_all("h1", class_="x-item-title__mainTitle")
    price = detail_soup.find_all("div", class_="x-price-primary")
    photos = detail_soup.find_all("div", class_="ux-image-carousel zoom img-transition-medium")#[0].findChildren()[0]
    
    
    description = detail_soup.find_all("iframe", id="desc_ifr")
    description_page = requests.get(description[0].get('src'))
    description_page_soup = BeautifulSoup(description_page.content, 'html.parser')
    description_message = description_page_soup.find_all("div", class_="x-item-description-child")

    category = detail_soup.find_all("a", class_="seo-breadcrumb-text")

    specification_title = detail_soup.find_all("dt", class_="ux-labels-values__labels")
    specification_value = detail_soup.find_all("dd", class_="ux-labels-values__values")
    
    category_text = ""
    for ix, cat in enumerate(category):
        if ix == 0:
            continue
        elif ix == 1:
            category_text += f"{cat.text} "
        else:
            category_text += f"> {cat.text} "
    
    category_text = category_text.strip()
    photos_list = []
    for ph in photos[0].findChildren():
        if ph.get('data-src'):
            photos_list.append(ph.get('data-src').replace("l140", "l1600"))

    product_specifics = []
    for index, value in enumerate(specification_title):
        product_specifics.append((value.text, specification_value[index].text))

    return [
        photos_list,
        title[0].text,
        description_message[0].text,
        price[0].text,
        category_text,
        target_url,
        product_specifics  # ix: 6
    ]

# BEGIN

file_path = 'input.xlsx'
urls = read_urls_from_excel(file_path)

print("Starting process")
# Create a new Workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

cols = ["Photos", "Title", "Description", "Price", "Category", "URL"]

product_detail = []
for url_active in urls:

    product_detail_req = get_detail_product(url_active)

    product_detail.append(product_detail_req)
    
    for nc in product_detail_req[6]:
        if nc[0] not in cols:
            cols.append(nc[0])

empty_spaces = ['-'] * len(cols)

worksheet.append(cols)

for index, value in enumerate(product_detail):
    values = empty_spaces

    links_str = ""
    for v in value[0]:
        if v == value[0][-1]:
            links_str += f"{v}"
        else:
            links_str += f"{v} \n"

    values[0] = links_str
    values[1] = value[1]
    values[2] = value[2]
    values[3] = value[3]
    values[4] = value[4]
    values[5] = value[5]
    
    for ix, det in enumerate(value[6]):
        if det[0] in cols:
            position = cols.index(det[0])
            values[position] = det[1]

    worksheet.append(values)

for ix, column_cells in enumerate(worksheet.columns):
    if ix == 0:
        worksheet.column_dimensions[column_cells[0].column_letter].width = 50
        continue
    
    if ix != 1:
        worksheet.row_dimensions[ix].height = 20

    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
workbook.save(f"output_{current_datetime}.xlsx")

print("End process")

# END
